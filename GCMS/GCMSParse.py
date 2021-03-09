# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

cr_umol = 1 #umol of Cr per reactor
nonane_mg = 1 #mg of nonane in each reactor, needs to be consistent across this set of experiments
volume_mL = 4 #initial volume of reaction solution
rxn_h = 1 #reaction time in hours
cr_l_ratio = 1 #Cr/L molar ratio

#calibration curve data--may need to modify this with new calibration curve values
d = {'compound_name':["1-hexene","methylcyclopentane","methylenecyclopentane","1-octene","C8isomers","PhCl","nonane","1-decene","C10isomers","1-dodecene","C12isomers","1-tetradecene","C14isomers","1-hexadecene","C16isomers","C18+"],
     'retention_time':[2.51,2.97,3.37,8.66,[8.75,8.95],9.33,11.12123213,12.57,[8.75,8.95],15.03,[8.75,8.95],16.97,[8.75,8.95],19.06,[8.75,8.95],[22.81,30.0]],
     'slope':[0.0047,0.0047,0.0047,0.0081,0.0081,0.0081,0.00893,0.0094,0.0094,0.0115,0.0115,0.0136,0.0136,0.0162,0.0162,0.0148], # slope is (area analyte/area 111.96 mM nonane/mM analyte)
     'intercept':[-0.0049,-0.0049,-0.0049,-0.009,-0.009,0,0,-0.0144,-0.0144,-0.0271,-0.0271,-0.0355,-0.0355,-0.0435,-0.0435,-0.0747]}
calib_date="[calibDate]"
offset=0.1 #fudge factor in minutes for peak identification

import pandas as pd
calib_curve = pd.DataFrame(data=d)

def index_to_letter(i):
    return chr(i+65)

def parse_GCMS(retention_times):
    """Parse all GC-MS OpenChrom Reports in the same directory as this python file
        
    Input: dictionary of retention times (floats, or a list of two floats defining a time range) for analytes, in minutes
    Output: pandas dataframe with GC-MS peak areas
        dataframe rows: each row corresponds to one .txt file
        dataframe columns: filename/index, 1-C6, C6isomers, 1-C8, C8isomers, PhCl, nonane, 1-C10, C10isomers, 1-C12, C12isomers, 1-C14, C14isomers, 1-C16, C16isomers, C18+
    """
    
    import os
    dir_path = os.path.dirname(os.path.realpath(__file__))
    print("parsing GC-MS directory: " + dir_path)
    columnNames = list(retention_times.keys())
    columnNames.insert(0,"filename")
    data =pd.DataFrame(columns=columnNames)
    for filename in os.listdir(dir_path):
        if filename.endswith(".txt"):
            print('parsing ' + filename)      
            file = open(dir_path + "/" + filename,'r')
            matchPeaks=0
            recordAreas=0
            peakNumbers=dict.fromkeys(retention_times.keys())
            areas=dict.fromkeys(retention_times.keys(),0)
            for line in file:
                #print(line)
                if line.find("RT (Min)") != -1:
                    matchPeaks=1
                elif matchPeaks == 1 and line == '\n':
                    matchPeaks = 0
                    #print(peakNumbers)
                    #print("finished assigning peaks")
                elif line.find("Integrated Area") != -1:
                    recordAreas=1
                elif recordAreas == 1 and line == '\n':
                    #print("finished recording peak areas")
                    #print(areas)
                    break
                elif matchPeaks == 1:
                    time = float(line.split()[1])
                    for key in retention_times:
                        t = retention_times[key]
                        #print(str(time)+ " : " + str(t))
                        if type(t) == float: #if specific analyte with single r.t.
                            if abs(time - t) < offset: 
                                peakNumbers[key] = line.split()[0]
                                #print("match!")
                                break
                        elif type(t) ==  list:  #elif range of r.t.'s
                            if time > t[0] and time < t[1]:
                                if peakNumbers[key] == None:
                                    peakNumbers[key]=[]
                                peakNumbers[key].append(line.split()[0])
                                #print(peakNumbers[key])
                                break
                elif recordAreas == 1:
                    peak = line.split()[0]
                    for key in peakNumbers:
                        p = peakNumbers[key]
                        if type(p) == str:
                            if peak == p:
                                areas[key] = float(line.split()[2])
                                break
                        elif type(p) == list:
                            if peak in p:
                                areas[key] += float(line.split()[2])
                                break
            areas['filename'] = filename
            #print(areas)                
            data = data.append(areas,ignore_index=True)
    return data

def GCMSarea_to_mass(areas):
    """Converts GC-MS areas to actual masses produced in reactions (averaged over multiple injections of one sample)
        
    Input: dataframe of GC-MS areas;
    Output: pandas dataframe with mgs produced in each reactor
        dataframe rows: each row corresponds to sample (perhaps the average of multiple GC-MS injections)
        dataframe columns: filename/index, 1-C6, C6isomers, 1-C8, C8isomers, PhCl, nonane, 1-C10, C10isomers, 1-C12, C12isomers, 1-C14, C14isomers, 1-C16, C16isomers, C18+
    """
 
    print("GCMS area to concentrations")
    masses = areas.copy()
    for i,row in areas.iterrows():
        #print("row: " + str(row))
        for j,analyte in calib_curve.iterrows():
            #print(analyte["slope"])
            if row[j+1] == 0:
                continue
            #print(analyte["intercept"])
            masses.at[i,analyte['compound_name']] = analyte["slope"]*row[j+1]+analyte["intercept"]
    print("now average conc. for multiple injections of same sample")
    return masses

def GCMSarea_to_excel(areas):
    """Converts GC-MS areas to Excel workbook with sheets listing reaction outcomes
    
    Global pre-reqs: nonane_mg, calib_curve    
    Input: dataframe of GC-MS areas
    Output: pandas dataframe with mgs produced in each reactor
        dataframe rows: each row corresponds to sample (perhaps the average of multiple GC-MS injections)
        dataframe columns: filename/index, 1-C6, C6isomers, 1-C8, C8isomers, PhCl, nonane, 1-C10, C10isomers, 1-C12, C12isomers, 1-C14, C14isomers, 1-C16, C16isomers, C18+
    """
    
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    import os
    print("writing areas to Excel: " + str(areas))
    wb = Workbook()
    ws0a = wb.active
    ws0a.title = "exp_parameters"
    ws0a.append(["file","Cr(umol)","nonane(mg)","initial volume (mL)", "reaction time(h)","Cr:L (molar ratio)", "PE (mg)"])
    for file in areas['filename']:
        ws0a.append([file,cr_umol,nonane_mg,volume_mL,rxn_h,cr_l_ratio])
    #enter Cr amt, nonane amt, GC-MS aliquot volume ratio, Cr:L ratio, reaction time, PE masses for each reactor
    #auto-fill in some fields
    
    ws0b = wb.create_sheet()
    ws0b.title = "calib_curve"
    for r in dataframe_to_rows(calib_curve,index=False,header=True):
        line = []
        for cell in r:
            if type(cell) == list:
                line.append(str(cell)) #cast typing necessary for arrays
            else:
                line.append(cell)
        ws0b.append(line)
    ws0b.append([])
    ws0b.append(["Units:","minutes, +/- " + str(offset),"mg/mL","TIC"])
    ws0b.append(["calibration data from: " + str(calib_date)])
    ws0b.append(["SL, 3/9/2021: calibration still needs tweaking"])
    
    ws0c = wb.create_sheet()
    ws0c.title = "math"
    ws0c.append(["**Known**"])
    ws0c.append(["best fit lines:","area(analyte) = m * conc(analyte) + b"])
    ws0c.append(["conc(nonane)_initial:","mg/mL from experimental setup","=exp_parameters!C2/exp_parameters!D2"])
    ws0c.append([])
    ws0c.append(["**Calculate**"])
    ws0c.append(["volume_final (mL)","volume_initial * conc(nonane)_final / conc(nonane)_initial","=exp_parameters!D2/calib_curve!C25*((GCMS_areas!H2-calib_curve!D8)/calib_curve!C8)"])
    ws0c.append(["conc(analyte)","from best fit line"])
    ws0c.append(["mg(analyte)","conc(analyte) * volume_final"])
    
    
    ws1 = wb.create_sheet()
    ws1.title = "GCMS_areas"
    for r in dataframe_to_rows(areas, index=False, header=True):
        ws1.append(r)
    
    #ws0a.append(["testing!"])
    
    ws2 = wb.create_sheet()
    ws2.title = "yield_mg"
    #write Excel equations to ws2, using pre-defined linear response factors, areas from ws1, and nonane amt
    masses = areas.copy()
    for i,row in areas.iterrows():
        #print("row: " + str(row))
        for j,analyte in calib_curve.iterrows():
            #print(analyte["slope"])
            if row[j+1] == 0:
                continue
            #print(analyte["intercept"])
            #masses.at[i,analyte['compound_name']] = analyte["slope"]*row[j+1]+analyte["intercept"]
            slope = "calib_curve!C" + str(j+2)
            area = "GCMS_areas!" + index_to_letter(j+2) + str(i+2)
            intercept = "calib_curve!D" + str(j+2)
            v_final = "math!C6"
            masses.loc[i,analyte['compound_name']] = "=(" + area + "-" + intercept + ")/" + slope + "*" + v_final
    for r in dataframe_to_rows(masses, index=False, header=True):
        ws2.append(r)
   
    ws2.append(["SL, 3/9/2021: calibration formula still needs tweaking"])
    
    ws3 = wb.create_sheet()
    ws3.title = "output"
    #write Excel equations to ws3, calculating activities and selecitivites from ws2, and input from ws: Cr amt, PE mass, reaction time
    
    ws4 = wb.create_sheet()
    ws4.title = "polished"
    #reformat useful data from ws3 into pretty form
    
    #for cell in ws['A'] + ws[1]:
    #    cell.style='Pandas'
    
    dir = os.path.dirname(os.path.realpath(__file__))    
    wb.save(os.path.basename(dir) + "_output.xlsx")
    
          
def write_output(df):
    #input: a pandas dataframe for GC-MS peak areas, calculated masses, and relative concentrations
    #output: Excel workbook with worksheets for (a) raw GC-MS peak areas (b) calculated mass yields (c) activities and wt% selectivities (requiring user input in the Excel table of Cr amount, nonane amount, reaction time, PE masses)
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    print("writing output Excel: data = " + str(df))
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "exp_parameters"
    #create ws0 to enter Cr amt, nonane amt, Cr:L ratio, reaction time, PE masses
    ws1 = wb.create_sheet()
    ws1.title = "GCMS_areas"
    
    for r in dataframe_to_rows(df, index=True, header=True):
        ws1.append(r)
    
    ws2 = wb.create_sheet()
    ws2.title = "yields_avg"
    #write Excel equations to ws2, using pre-defined linear response factors, areas from ws1, and nonane amt
    
    ws3 = wb.create_sheet()
    ws3.title = "output"
    #write Excel equations to ws3, calculating activities and selecitivites from ws2, and input from ws: Cr amt, PE mass, reaction time
    
    ws4 = wb.create()
    ws4.title = "polished"
    #reformat useful data from ws3 into pretty form
    
    #for cell in ws['A'] + ws[1]:
    #    cell.style='Pandas'
        
    wb.save("output.xlsx")
    


retention_times = {calib_curve['compound_name'][i] : calib_curve['retention_time'][i] for i in range(0,len(calib_curve['compound_name']))}
raw_data = parse_GCMS(retention_times)
print("****Raw Data****")
print(raw_data)
#mass_data = GCMSarea_to_mass(raw_data)
#print("****Mass Data****")
#print(mass_data)
GCMSarea_to_excel(raw_data)


"""
experiments=["AOT-SL-130-1","AOT-SL-130-2","AOT-SL-130-3","AOT-SL-130-4"]
columns=["experiment","1-C6","H3C5a","H2C5e","1-C8", "C8", "PhCl", "C9", "1-C10","1-C12_mg","1-C14_mg","1-C16+_mg","PE_mg","TotalProducts_mg","Activity_g/gCr/h",
         "1-C6_wt%","H3C5a_wt%","H2C5e_wt%","1-C8_wt%","1-C10_wt%","1-C12_wt%","1-C14_wt%","1-C16+_wt%","PE_mg"]
retentionTimes={
    "1-C6":1.1,
    "methylcyclopentane":1.1,
    "methylenecyclopentane":1.3,
    "1-C8":1.4,
    "PhCl":1.5,
    "C9":1.6,
    "1-C10":1.7,
    "1-C12":1.8,
    "1-C14":1.9
    } #seconds(?) these may change with column aging

offset=0.5 #fudge time to account for a peak


linearResponseFactors={
    "1-C6":1.1,
    "methylcyclopentane":1.1,
    "methylenecyclopentane":1.3,
    "1-C8":1.4,
    "PhCl":1.5,
    "C9":1.6,
    "1-C10":1.7,
    "1-C12":1.8,
    "1-C14":1.9
    } #these may change as MS ages or is retuned
data = pd.DataFrame(index=experiments, columns=retentionTimes.keys())
data = data.fillna(0)

#write_output(data)
"""
